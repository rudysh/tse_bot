import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class ExcelService:
    """Lee y actualiza el mismo archivo Excel de entrada."""

    relleno_rojo_suave = PatternFill(fill_type="solid", fgColor="FCE4D6")
    relleno_amarillo_suave = PatternFill(fill_type="solid", fgColor="FFF2CC")

    def __init__(self, ruta_archivo: Path) -> None:
        self.ruta_archivo = ruta_archivo
        self.workbook = load_workbook(ruta_archivo)
        self.sheet = self.workbook.active
        self.fila_inicio = self._detectar_fila_inicio()
        self._crear_encabezados_si_corresponde()

    def _detectar_fila_inicio(self) -> int:
        primera_cedula = self.limpiar_cedula(self.sheet.cell(row=1, column=1).value)
        return 2 if not primera_cedula.isdigit() else 1

    def _crear_encabezados_si_corresponde(self) -> None:
        if self.fila_inicio != 2:
            return

        encabezados = {
            2: "Nombre completo",
            3: "Estado",
            4: "Fecha y hora de consulta",
        }
        for columna, titulo in encabezados.items():
            if self.sheet.cell(row=1, column=columna).value is None:
                self.sheet.cell(row=1, column=columna).value = titulo

    @staticmethod
    def limpiar_cedula(valor: object) -> str:
        if valor is None:
            return ""
        return re.sub(r"[\s-]+", "", str(valor).strip())

    def iterar_cedulas(self) -> Iterable[tuple[int, str]]:
        for fila in range(self.fila_inicio, self.sheet.max_row + 1):
            cedula = self.limpiar_cedula(self.sheet.cell(row=fila, column=1).value)
            yield fila, cedula

    def total_consultas(self) -> int:
        return max(self.sheet.max_row - self.fila_inicio + 1, 0)

    def escribir_resultado(self, fila: int, nombre: str, estado: str, fecha_hora: str) -> None:
        self.sheet.cell(row=fila, column=2).value = nombre
        self.sheet.cell(row=fila, column=3).value = estado
        self.sheet.cell(row=fila, column=4).value = fecha_hora
        self._aplicar_color_fila(fila, estado)

    def guardar(self) -> None:
        self.workbook.save(self.ruta_archivo)

    def _aplicar_color_fila(self, fila: int, estado: str) -> None:
        relleno = None
        if estado == "Cédula inválida":
            relleno = self.relleno_rojo_suave
        elif estado == "No encontrado":
            relleno = self.relleno_amarillo_suave

        if relleno is None:
            return

        for columna in range(1, self.sheet.max_column + 1):
            self.sheet.cell(row=fila, column=columna).fill = relleno
